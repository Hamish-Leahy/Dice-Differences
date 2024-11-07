import tkinter as tk
from tkinter import ttk, messagebox
import random
import pandas as pd
from datetime import datetime
import threading
import numpy as np
from itertools import product
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class DiceDifferencesAnalyzer:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Dice Differences Game Analyzer")
        self.root.geometry("800x600")
        
        self.running = False
        self.results = []
        self.dice_rolls = []
        self.all_results = []
        self.display_delay = int(1000 / 90)  # Set the delay for 90x real-time speed
        
        self.setup_gui()

    def setup_gui(self):
        # Strategy input frame
        strategy_frame = ttk.LabelFrame(self.root, text="Counter Placement Strategy (0-6 counters for 6 positions)", padding="10")
        strategy_frame.pack(fill="x", padx=10, pady=5)
        
        self.counter_entries = []
        for i in range(6):  # 6 positions (0 to 5) for differences
            frame = ttk.Frame(strategy_frame)
            frame.pack(fill="x", pady=2)
            ttk.Label(frame, text=f"Difference {i}: ").pack(side="left")
            spinbox = ttk.Spinbox(frame, from_=0, to=6, width=5)
            spinbox.set(0)
            spinbox.pack(side="left")
            self.counter_entries.append(spinbox)
        
        # Simulation controls
        control_frame = ttk.LabelFrame(self.root, text="Simulation Controls", padding="10")
        control_frame.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(control_frame, text="Number of games:").pack(side="left")
        self.num_games = ttk.Entry(control_frame, width=10)
        self.num_games.insert(0, "1000")
        self.num_games.pack(side="left", padx=5)
        
        self.start_button = ttk.Button(control_frame, text="Start Manual Simulation", command=self.start_manual_simulation)
        self.start_button.pack(side="left", padx=5)
        
        self.auto_button = ttk.Button(control_frame, text="Start Auto Mode", command=self.start_auto_simulation)
        self.auto_button.pack(side="left", padx=5)
        
        self.stop_button = ttk.Button(control_frame, text="Stop", command=self.stop_simulation, state="disabled")
        self.stop_button.pack(side="left", padx=5)
        
        # Results display
        results_frame = ttk.LabelFrame(self.root, text="Results", padding="10")
        results_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.text_output = tk.Text(results_frame, height=20)
        self.text_output.pack(fill="both", expand=True)

    def display_results(self, game_rolls):
        """Display game results with a fast-forward effect."""
        self.text_output.delete("1.0", tk.END)  # Clear previous results

        def update_display(index=0):
            if index < len(game_rolls):
                die1, die2, diff, success = game_rolls[index]
                self.text_output.insert(tk.END, f"Move {index + 1}: Die1={die1}, Die2={die2}, "
                                                f"Difference={diff} -> {'Counter removed' if success else 'No counter removed'}\n")
                self.text_output.see(tk.END)
                # Schedule the next line to display after the delay
                self.root.after(self.display_delay, update_display, index + 1)

        update_display()

    def get_strategy(self):
        """Retrieve the manually configured strategy."""
        strategy = {}
        total_counters = 0
        
        for i, entry in enumerate(self.counter_entries):
            try:
                count = int(entry.get())
                if count < 0 or count > 6:
                    raise ValueError
                if count > 0:
                    strategy[i] = count
                total_counters += count
            except ValueError:
                messagebox.showerror("Error", f"Invalid input for difference {i}")
                return None
        
        if total_counters != 6:
            messagebox.showerror("Error", "Total counters must equal 6")
            return None
            
        return strategy
    
    def play_single_game(self, strategy, output=False):
        """Simulate a single game based on a given strategy."""
        counters = []
        for diff, count in strategy.items():
            counters.extend([diff] * count)
        
        moves = 0
        game_rolls = []
        
        while counters:
            moves += 1
            die1 = random.randint(1, 6)
            die2 = random.randint(1, 6)
            diff = abs(die1 - die2)
            success = diff in counters
            
            if output:  # Only output in manual mode
                game_rolls.append((die1, die2, diff, success))
                
            if success:
                counters.remove(diff)
                
        return moves, game_rolls

    def simulate_games(self, strategy, num_games, output=False):
        """Simulate multiple games for a given strategy."""
        self.results = []
        self.dice_rolls = []
        
        for i in range(num_games):
            moves, game_rolls = self.play_single_game(strategy, output)
            self.results.append(moves)
            self.dice_rolls.append(game_rolls)
            if output:
                self.display_results(game_rolls)

    def save_manual_results(self, strategy):
        """Save manual simulation results to an Excel file."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dice_differences_manual_results_{timestamp}.xlsx"
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Detailed Rolls
            all_rolls = []
            for game_num, game_rolls in enumerate(self.dice_rolls, 1):
                for move_num, (die1, die2, diff, success) in enumerate(game_rolls, 1):
                    all_rolls.append({
                        'Game': game_num,
                        'Move': move_num,
                        'Die 1': die1,
                        'Die 2': die2,
                        'Difference': diff,
                        'Counter Removed': 'Yes' if success else 'No'
                    })
            pd.DataFrame(all_rolls).to_excel(writer, sheet_name='Detailed Rolls', index=False)

            # Summary
            summary_data = {
                'Game': list(range(1, len(self.results) + 1)),
                'Total Moves': self.results
            }
            
            # Calculate additional metrics for a single summary row
            avg_moves = np.mean(self.results)
            min_moves = min(self.results)
            max_moves = max(self.results)
            
            summary_stats = pd.DataFrame([{
                'Average Moves': avg_moves,
                'Min Moves': min_moves,
                'Max Moves': max_moves
            }])
            
            # Write main summary and summary stats to Excel
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', index=False)
            summary_stats.to_excel(writer, sheet_name='Summary', index=False, startrow=len(self.results) + 2)

    def start_manual_simulation(self):
        """Start a manual simulation with user-defined strategy."""
        strategy = self.get_strategy()
        if not strategy:
            return

        num_games = int(self.num_games.get())
        self.running = True
        self.start_button.config(state="disabled")
        self.stop_button.config(state="normal")

        self.simulate_games(strategy, num_games, output=True)
        self.save_manual_results(strategy)

        self.start_button.config(state="normal")
        self.stop_button.config(state="disabled")
        self.running = False
    
    def analyze_all_results(self):
        """Analyze results and save sorted final evaluation for auto mode."""
        evaluation_data = []

        for strategy, results in self.all_results:
            avg_moves = np.mean(results)
            min_moves = min(results)
            max_moves = max(results)
            evaluation_data.append({
                'Strategy': strategy,
                'Average Moves': avg_moves,
                'Min Moves': min_moves,
                'Max Moves': max_moves
            })

        # Convert evaluation data to a DataFrame and sort by Average Moves
        evaluation_df = pd.DataFrame(evaluation_data).sort_values(by='Average Moves')
        
        # Save the evaluation to an Excel file
        evaluation_filename = f"dice_differences_final_evaluation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        with pd.ExcelWriter(evaluation_filename, engine='openpyxl') as writer:
            evaluation_df.to_excel(writer, sheet_name='Final Evaluation', index=False)
        
        # Highlight the best strategy (lowest Average Moves)
        workbook = load_workbook(evaluation_filename)
        sheet = workbook['Final Evaluation']
        best_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Apply highlight to the best strategy row (first data row)
        for cell in sheet[2]:  # Assuming headers are in the first row
            cell.fill = best_fill
            
        workbook.save(evaluation_filename)

    def auto_simulate_all_combinations(self):
        """Run auto simulation on all valid counter configurations."""
        self.all_results = []
        num_games = int(self.num_games.get())
        
        for combination in product(range(7), repeat=6):  # 6 counters across 6 positions (0 to 5)
            if sum(combination) == 6:
                strategy = {i: count for i, count in enumerate(combination) if count > 0}
                self.simulate_games(strategy, num_games, output=False)
                self.all_results.append((strategy, self.results.copy()))
        
        self.analyze_all_results()

    def start_auto_simulation(self):
        """Start auto mode simulation to test all possible strategies."""
        self.running = True
        self.start_button.config(state="disabled")
        self.auto_button.config(state="disabled")
        self.stop_button.config(state="normal")
        
        auto_thread = threading.Thread(target=self.auto_simulate_all_combinations)
        auto_thread.start()
    
    def stop_simulation(self):
        """Stop any ongoing simulation."""
        self.running = False
    
    def run(self):
        """Run the application."""
        self.root.mainloop()

if __name__ == "__main__":
    app = DiceDifferencesAnalyzer()
    app.run()
